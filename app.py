from flask import Flask, render_template, request, jsonify, session, redirect, send_file, Response
import sqlite3, hashlib, json, csv, io
from datetime import datetime
from functools import wraps

app = Flask(__name__)
app.secret_key = "agro_nuts_2026_secret_key_change_in_production"
DB = "agro_nuts_v2.db"

# ── DB ─────────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db(); c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL, password TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'viewer',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")

    c.execute("""CREATE TABLE IF NOT EXISTS suppliers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
    # Remove old budget column — budget is now SUM of allocations
    # Keep column for backwards compat but we won't use it for calculations
    existing_sup = [r[1] for r in c.execute("PRAGMA table_info(suppliers)").fetchall()]
    if "budget" not in existing_sup:
        c.execute("ALTER TABLE suppliers ADD COLUMN budget REAL DEFAULT 0")

    # ── NEW: Allocations table ─────────────────────────────────────────────────
    c.execute("""CREATE TABLE IF NOT EXISTS allocations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        supplier_id INTEGER NOT NULL,
        date TEXT NOT NULL,
        amount REAL NOT NULL DEFAULT 0,
        notes TEXT DEFAULT '',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id))""")

    c.execute("""CREATE TABLE IF NOT EXISTS deliveries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL, truck TEXT NOT NULL,
        bags INTEGER DEFAULT 0, gross_weight REAL DEFAULT 0, net_weight REAL DEFAULT 0,
        lbs REAL DEFAULT 0, bags_after_dry INTEGER DEFAULT 0,
        quality_after_dry TEXT DEFAULT '', count_cashew REAL DEFAULT 0,
        fd TEXT DEFAULT '', gk TEXT DEFAULT '',
        price REAL DEFAULT 0, amount REAL DEFAULT 0, supplier_id INTEGER,
        sold_outside INTEGER DEFAULT 0, sold_price REAL DEFAULT 0, sold_amount REAL DEFAULT 0,
        notes TEXT DEFAULT '', created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (supplier_id) REFERENCES suppliers(id))""")

    # Migrate deliveries columns
    existing_del = [r[1] for r in c.execute("PRAGMA table_info(deliveries)").fetchall()]
    for col, defn in [("lbs","REAL DEFAULT 0"),("bags_after_dry","INTEGER DEFAULT 0"),
                      ("quality_after_dry","TEXT DEFAULT ''"),("count_cashew","REAL DEFAULT 0"),
                      ("fd","TEXT DEFAULT ''"),("gk","TEXT DEFAULT ''")]:
        if col not in existing_del:
            c.execute(f"ALTER TABLE deliveries ADD COLUMN {col} {defn}")

    # Seed users
    c.execute("INSERT OR IGNORE INTO users (username,password,role) VALUES (?,?,?)",
              ("admin",  hashlib.sha256("admin123".encode()).hexdigest(),  "admin"))
    c.execute("INSERT OR IGNORE INTO users (username,password,role) VALUES (?,?,?)",
              ("viewer", hashlib.sha256("viewer123".encode()).hexdigest(), "viewer"))

    # Seed suppliers (no budget — will be added via allocations)
    defaults = ["MALAN BAMBADINCA","USMANE BUBA","SIDI EBAYA QUEBO",
                "LAMARANA BALDE QUEBO","IDUMO","MAMADU LAMARANA BA GILAGE",
                "MAJEEDI","NIMAD SIDABY"]
    for name in defaults:
        c.execute("INSERT OR IGNORE INTO suppliers (name) VALUES (?)", (name,))

    # Migrate: if suppliers had budgets but no allocations, create a seed allocation
    for row in c.execute("SELECT id, budget FROM suppliers WHERE budget > 0").fetchall():
        exists = c.execute("SELECT COUNT(*) FROM allocations WHERE supplier_id=?", (row[0],)).fetchone()[0]
        if exists == 0:
            c.execute("INSERT INTO allocations (supplier_id, date, amount, notes) VALUES (?,?,?,?)",
                      (row[0], datetime.today().strftime("%d-%m-%Y"), row[0], "Initial budget (migrated)"))

    conn.commit(); conn.close()

# ── Auth ───────────────────────────────────────────────────────────────────────
def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def d(*a, **k):
        if "user_id" not in session: return jsonify({"error": "unauthorized"}), 401
        return f(*a, **k)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a, **k):
        if session.get("role") != "admin": return jsonify({"error": "admin only"}), 403
        return f(*a, **k)
    return d

# ── Stats ──────────────────────────────────────────────────────────────────────
def supplier_stats(sid, conn):
    c = conn.cursor()
    # Total allocated = sum of all allocation payments
    c.execute("SELECT COALESCE(SUM(amount),0) FROM allocations WHERE supplier_id=?", (sid,))
    total_allocated = c.fetchone()[0]
    # Amount used = sum of warehouse amounts (not sold outside)
    c.execute("SELECT COALESCE(SUM(amount),0), COUNT(*) FROM deliveries WHERE supplier_id=? AND sold_outside=0", (sid,))
    r = c.fetchone(); used, count_in = r[0], r[1]
    # Sold outside
    c.execute("SELECT COALESCE(SUM(sold_amount),0), COUNT(*) FROM deliveries WHERE supplier_id=? AND sold_outside=1", (sid,))
    r2 = c.fetchone(); sold_amt, count_out = r2[0], r2[1]
    balance = total_allocated - used
    pct = (balance / total_allocated * 100) if total_allocated > 0 else -1
    status = "OVER BUDGET" if pct < 0 else "CRITICAL" if pct < 10 else "LOW" if pct < 30 else "OK"
    return {"total_allocated": total_allocated, "used": used, "balance": balance,
            "pct": round(pct, 1), "status": status,
            "count_in": count_in, "count_out": count_out, "sold_amount": sold_amt}

def parse_delivery(d):
    net = float(d.get("net_weight") or 0); price = float(d.get("price") or 0)
    sold = 1 if d.get("sold_outside") else 0; sp = float(d.get("sold_price") or 0)
    return {"date": d.get("date", datetime.today().strftime("%d-%m-%Y")),
            "truck": d.get("truck", ""), "bags": int(d.get("bags") or 0),
            "gross_weight": float(d.get("gross_weight") or 0), "net_weight": net,
            "lbs": float(d.get("lbs") or 0), "bags_after_dry": int(d.get("bags_after_dry") or 0),
            "quality_after_dry": str(d.get("quality_after_dry") or ""),
            "count_cashew": float(d.get("count_cashew") or 0),
            "fd": str(d.get("fd") or ""), "gk": str(d.get("gk") or ""),
            "price": price, "amount": net * price, "supplier_id": d.get("supplier_id"),
            "sold_outside": sold, "sold_price": sp,
            "sold_amount": net * sp if sold else 0, "notes": str(d.get("notes") or "")}

# ── Pages ──────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    if "user_id" not in session: return render_template("login.html")
    return render_template("app.html", username=session.get("username"), role=session.get("role"))

@app.route("/login", methods=["POST"])
def login():
    d = request.json; conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM users WHERE username=? AND password=?",
              (d.get("username","").strip(), hash_pw(d.get("password",""))))
    user = c.fetchone(); conn.close()
    if user:
        session.update({"user_id": user["id"], "username": user["username"], "role": user["role"]})
        return jsonify({"success": True, "role": user["role"], "username": user["username"]})
    return jsonify({"success": False, "error": "Invalid credentials"}), 401

@app.route("/logout")
def logout(): session.clear(); return redirect("/")

# ── Dashboard ──────────────────────────────────────────────────────────────────
@app.route("/api/dashboard")
@login_required
def api_dashboard():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM suppliers ORDER BY name")
    sups = [dict(r) for r in c.fetchall()]
    result = []; totals = {"total_allocated":0,"used":0,"balance":0,"deliveries":0,"sold_out":0,"sold_amount":0}
    for s in sups:
        st = supplier_stats(s["id"], conn)
        totals["total_allocated"] += st["total_allocated"]
        totals["used"]            += st["used"]
        totals["balance"]         += st["balance"]
        totals["deliveries"]      += st["count_in"]
        totals["sold_out"]        += st["count_out"]
        totals["sold_amount"]     += st["sold_amount"]
        result.append({**s, **st})
    conn.close()
    return jsonify({"suppliers": result, "totals": totals})

# ── Allocations ────────────────────────────────────────────────────────────────
@app.route("/api/allocations")
@login_required
def api_allocations():
    sid = request.args.get("supplier_id")
    conn = get_db(); c = conn.cursor()
    if sid:
        c.execute("""SELECT a.*, s.name as supplier_name FROM allocations a
                     JOIN suppliers s ON a.supplier_id=s.id
                     WHERE a.supplier_id=? ORDER BY a.date DESC, a.id DESC""", (sid,))
    else:
        c.execute("""SELECT a.*, s.name as supplier_name FROM allocations a
                     JOIN suppliers s ON a.supplier_id=s.id
                     ORDER BY a.date DESC, a.id DESC""")
    rows = [dict(r) for r in c.fetchall()]; conn.close()
    return jsonify(rows)

@app.route("/api/allocations", methods=["POST"])
@login_required
@admin_required
def add_allocation():
    d = request.json
    supplier_id = d.get("supplier_id")
    date        = d.get("date", datetime.today().strftime("%d-%m-%Y"))
    amount      = float(d.get("amount") or 0)
    notes       = str(d.get("notes") or "")
    if not supplier_id or amount <= 0:
        return jsonify({"success": False, "error": "Supplier and amount required"}), 400
    conn = get_db(); c = conn.cursor()
    c.execute("INSERT INTO allocations (supplier_id, date, amount, notes) VALUES (?,?,?,?)",
              (supplier_id, date, amount, notes))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"success": True, "id": new_id})

@app.route("/api/allocations/<int:aid>", methods=["PUT"])
@login_required
@admin_required
def update_allocation(aid):
    d = request.json; conn = get_db()
    conn.execute("UPDATE allocations SET date=?, amount=?, notes=? WHERE id=?",
                 (d.get("date"), float(d.get("amount") or 0), d.get("notes",""), aid))
    conn.commit(); conn.close()
    return jsonify({"success": True})

@app.route("/api/allocations/<int:aid>", methods=["DELETE"])
@login_required
@admin_required
def delete_allocation(aid):
    conn = get_db()
    conn.execute("DELETE FROM allocations WHERE id=?", (aid,))
    conn.commit(); conn.close()
    return jsonify({"success": True})

@app.route("/api/suppliers/<int:sid>/allocations")
@login_required
def supplier_allocations(sid):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM allocations WHERE supplier_id=? ORDER BY date DESC, id DESC", (sid,))
    rows = [dict(r) for r in c.fetchall()]
    running = 0
    for r in reversed(rows): running += r["amount"]; r["running_total"] = running
    rows.reverse()
    conn.close(); return jsonify(rows)

# ── Deliveries ─────────────────────────────────────────────────────────────────
@app.route("/api/deliveries")
@login_required
def api_deliveries():
    sid = request.args.get("supplier_id"); sold = request.args.get("sold")
    search = request.args.get("search","").strip()
    conn = get_db(); c = conn.cursor()
    q = "SELECT d.*,s.name as supplier_name FROM deliveries d JOIN suppliers s ON d.supplier_id=s.id WHERE 1=1"
    p = []
    if sid:  q += " AND d.supplier_id=?"; p.append(sid)
    if sold is not None and sold != "": q += " AND d.sold_outside=?"; p.append(sold)
    if search: q += " AND (d.truck LIKE ? OR s.name LIKE ?)"; p += [f"%{search}%"]*2
    q += " ORDER BY d.date DESC, d.id DESC"
    c.execute(q, p); rows = [dict(r) for r in c.fetchall()]; conn.close()
    return jsonify(rows)

@app.route("/api/deliveries/<int:did>")
@login_required
def get_delivery(did):
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT d.*,s.name as supplier_name FROM deliveries d JOIN suppliers s ON d.supplier_id=s.id WHERE d.id=?", (did,))
    row = c.fetchone(); conn.close()
    return jsonify(dict(row)) if row else (jsonify({"error":"not found"}), 404)

@app.route("/api/deliveries", methods=["POST"])
@login_required
@admin_required
def add_delivery():
    f = parse_delivery(request.json); conn = get_db(); c = conn.cursor()
    c.execute("""INSERT INTO deliveries (date,truck,bags,gross_weight,net_weight,lbs,bags_after_dry,
        quality_after_dry,count_cashew,fd,gk,price,amount,supplier_id,sold_outside,sold_price,sold_amount,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        tuple(f[k] for k in ["date","truck","bags","gross_weight","net_weight","lbs","bags_after_dry",
                              "quality_after_dry","count_cashew","fd","gk","price","amount","supplier_id",
                              "sold_outside","sold_price","sold_amount","notes"]))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"success": True, "id": new_id, "amount": f["amount"]})

@app.route("/api/deliveries/<int:did>", methods=["PUT"])
@login_required
@admin_required
def update_delivery(did):
    f = parse_delivery(request.json); conn = get_db()
    conn.execute("""UPDATE deliveries SET date=?,truck=?,bags=?,gross_weight=?,net_weight=?,lbs=?,
        bags_after_dry=?,quality_after_dry=?,count_cashew=?,fd=?,gk=?,price=?,amount=?,supplier_id=?,
        sold_outside=?,sold_price=?,sold_amount=?,notes=? WHERE id=?""",
        tuple(f[k] for k in ["date","truck","bags","gross_weight","net_weight","lbs","bags_after_dry",
                              "quality_after_dry","count_cashew","fd","gk","price","amount","supplier_id",
                              "sold_outside","sold_price","sold_amount","notes"]) + (did,))
    conn.commit(); conn.close(); return jsonify({"success": True})

@app.route("/api/deliveries/<int:did>", methods=["DELETE"])
@login_required
@admin_required
def delete_delivery(did):
    conn = get_db(); conn.execute("DELETE FROM deliveries WHERE id=?", (did,))
    conn.commit(); conn.close(); return jsonify({"success": True})

# ── Suppliers ──────────────────────────────────────────────────────────────────
@app.route("/api/suppliers")
@login_required
def api_suppliers():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM suppliers ORDER BY name")
    rows = [dict(r) for r in c.fetchall()]; conn.close(); return jsonify(rows)

@app.route("/api/suppliers", methods=["POST"])
@login_required
@admin_required
def add_supplier():
    d = request.json; name = d.get("name","").strip().upper()
    if not name: return jsonify({"success": False, "error": "Name required"}), 400
    conn = get_db()
    try:
        conn.execute("INSERT INTO suppliers (name) VALUES (?)", (name,))
        conn.commit(); sid = conn.execute("SELECT id FROM suppliers WHERE name=?", (name,)).fetchone()[0]
        # Add first allocation if provided
        if d.get("initial_amount") and float(d["initial_amount"]) > 0:
            conn.execute("INSERT INTO allocations (supplier_id,date,amount,notes) VALUES (?,?,?,?)",
                         (sid, d.get("initial_date", datetime.today().strftime("%d-%m-%Y")),
                          float(d["initial_amount"]), d.get("initial_notes","Initial allocation")))
            conn.commit()
        conn.close(); return jsonify({"success": True})
    except sqlite3.IntegrityError:
        conn.close(); return jsonify({"success": False, "error": "Already exists"}), 400

@app.route("/api/suppliers/<int:sid>", methods=["PUT"])
@login_required
@admin_required
def update_supplier(sid):
    d = request.json; conn = get_db()
    if "name" in d: conn.execute("UPDATE suppliers SET name=? WHERE id=?", (d["name"].strip().upper(), sid))
    conn.commit(); conn.close(); return jsonify({"success": True})

@app.route("/api/suppliers/<int:sid>", methods=["DELETE"])
@login_required
@admin_required
def delete_supplier(sid):
    force = request.args.get("force") == "1"
    conn = get_db(); c = conn.cursor()
    # Check if supplier has deliveries or allocations
    c.execute("SELECT COUNT(*) FROM deliveries WHERE supplier_id=?", (sid,))
    del_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM allocations WHERE supplier_id=?", (sid,))
    alloc_count = c.fetchone()[0]
    if (del_count > 0 or alloc_count > 0) and not force:
        conn.close()
        return jsonify({"success": False, "has_data": True,
                        "del_count": del_count, "alloc_count": alloc_count,
                        "error": f"Supplier has {del_count} deliveries and {alloc_count} allocations."}), 409
    # Delete everything related
    conn.execute("DELETE FROM deliveries WHERE supplier_id=?", (sid,))
    conn.execute("DELETE FROM allocations WHERE supplier_id=?", (sid,))
    conn.execute("DELETE FROM suppliers WHERE id=?", (sid,))
    conn.commit(); conn.close()
    return jsonify({"success": True})

# ── Sold summary ───────────────────────────────────────────────────────────────
@app.route("/api/sold-summary")
@login_required
def sold_summary():
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT s.name, s.id,
        COUNT(CASE WHEN d.sold_outside=0 THEN 1 END) as kept,
        COUNT(CASE WHEN d.sold_outside=1 THEN 1 END) as sold_out, COUNT(*) as total,
        COALESCE(SUM(CASE WHEN d.sold_outside=0 THEN d.net_weight END),0) as kept_weight,
        COALESCE(SUM(CASE WHEN d.sold_outside=1 THEN d.net_weight END),0) as sold_weight,
        COALESCE(SUM(CASE WHEN d.sold_outside=1 THEN d.sold_amount END),0) as sold_revenue,
        COALESCE(SUM(d.net_weight),0) as total_weight
        FROM suppliers s LEFT JOIN deliveries d ON s.id=d.supplier_id
        GROUP BY s.id ORDER BY s.name""")
    rows = [dict(r) for r in c.fetchall()]; conn.close(); return jsonify(rows)

# ── Users ──────────────────────────────────────────────────────────────────────
@app.route("/api/users")
@login_required
@admin_required
def api_users():
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id,username,role,created_at FROM users ORDER BY role,username")
    rows = [dict(r) for r in c.fetchall()]; conn.close(); return jsonify(rows)

@app.route("/api/users", methods=["POST"])
@login_required
@admin_required
def add_user():
    d = request.json; u = d.get("username","").strip(); pw = d.get("password","").strip()
    if not u or not pw: return jsonify({"success": False, "error": "Username and password required"}), 400
    conn = get_db()
    try:
        conn.execute("INSERT INTO users (username,password,role) VALUES (?,?,?)",
                     (u, hash_pw(pw), d.get("role","viewer")))
        conn.commit(); conn.close(); return jsonify({"success": True})
    except sqlite3.IntegrityError:
        conn.close(); return jsonify({"success": False, "error": "Username already exists"}), 400

@app.route("/api/users/<int:uid>", methods=["DELETE"])
@login_required
@admin_required
def delete_user(uid):
    if uid == session.get("user_id"): return jsonify({"success": False, "error": "Cannot delete yourself"}), 400
    conn = get_db(); conn.execute("DELETE FROM users WHERE id=?", (uid,))
    conn.commit(); conn.close(); return jsonify({"success": True})

# ── Export ─────────────────────────────────────────────────────────────────────
@app.route("/api/export/<fmt>")
@login_required
def export(fmt):
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT d.date,d.truck,d.bags,d.gross_weight,d.net_weight,d.lbs,d.bags_after_dry,
        d.quality_after_dry,d.count_cashew,d.fd,d.gk,d.price,d.amount,s.name as supplier,
        CASE WHEN d.sold_outside=1 THEN 'YES' ELSE 'NO' END as sold_outside,
        d.sold_price,d.sold_amount,d.notes
        FROM deliveries d JOIN suppliers s ON d.supplier_id=s.id ORDER BY d.date DESC""")
    rows = [dict(r) for r in c.fetchall()]
    # Also export allocations
    c.execute("""SELECT a.date, s.name as supplier, a.amount, a.notes
                 FROM allocations a JOIN suppliers s ON a.supplier_id=s.id ORDER BY a.date DESC""")
    alloc_rows = [dict(r) for r in c.fetchall()]
    conn.close()

    hdrs = ["Date","Truck","Bags","Gross Weight","Net Weight","LBS","Bags After Dry",
            "Quality After Dry","Count Cashew","F/D","GK","Price","Amount","Supplier",
            "Sold Outside","Sold Price","Sold Amount","Notes"]
    alloc_hdrs = ["Date","Supplier","Amount Allocated","Notes"]

    if fmt == "json":
        return Response(json.dumps({"deliveries": rows, "allocations": alloc_rows}, indent=2),
                        mimetype="application/json",
                        headers={"Content-Disposition":"attachment;filename=agro_nuts_export.json"})
    elif fmt == "csv":
        si = io.StringIO(); w = csv.writer(si)
        w.writerow(["=== DELIVERIES ==="]); w.writerow(hdrs)
        for r in rows: w.writerow(list(r.values()))
        w.writerow([]); w.writerow(["=== ALLOCATIONS ==="]); w.writerow(alloc_hdrs)
        for r in alloc_rows: w.writerow(list(r.values()))
        return Response(si.getvalue(), mimetype="text/csv",
                        headers={"Content-Disposition":"attachment;filename=agro_nuts_export.csv"})
    elif fmt == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            # Sheet 1: Deliveries
            ws = wb.active; ws.title = "Deliveries"
            for i, h in enumerate(hdrs, 1):
                cell = ws.cell(1, i, h)
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                cell.fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(i)].width = max(len(h)+4, 12)
            for ri, row in enumerate(rows, 2):
                for ci, v in enumerate(row.values(), 1):
                    ws.cell(ri, ci, v).font = Font(name="Arial", size=10)
                    if ri % 2 == 0: ws.cell(ri, ci).fill = PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB")
            # Sheet 2: Allocations
            wa = wb.create_sheet("Allocations")
            for i, h in enumerate(alloc_hdrs, 1):
                cell = wa.cell(1, i, h)
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                cell.fill = PatternFill("solid", start_color="27AE60", end_color="27AE60")
                cell.alignment = Alignment(horizontal="center")
                wa.column_dimensions[get_column_letter(i)].width = max(len(h)+4, 16)
            for ri, row in enumerate(alloc_rows, 2):
                for ci, v in enumerate(row.values(), 1):
                    wa.cell(ri, ci, v).font = Font(name="Arial", size=10)
                    if ri % 2 == 0: wa.cell(ri, ci).fill = PatternFill("solid", start_color="D4EDDA", end_color="D4EDDA")
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             as_attachment=True, download_name="agro_nuts_export.xlsx")
        except ImportError:
            return jsonify({"error": "openpyxl not installed"}), 500
    return jsonify({"error": "unknown format"}), 400

@app.route("/api/me")
def me():
    if "user_id" in session: return jsonify({"username": session["username"], "role": session["role"]})
    return jsonify({"error": "not logged in"}), 401

if __name__ == "__main__":
    init_db()
    print("\n✅  AGRO NUTS 2026 v4 — http://127.0.0.1:5000")
    print("   admin/admin123  |  viewer/viewer123\n")
    app.run(debug=True, port=5000)
